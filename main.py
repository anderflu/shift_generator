import pandas as pd
import numpy as np
from collections import defaultdict, Counter
import random
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

#Path til vaktliste
vaktliste_file_path = "test_data/Vaktønsker uke 41 & 42.xlsx"
number_of_shifts_week_1 = 14
number_of_shifts_week_2 = 14


# Load the sheet into a dataframe
sheet_name = 0
df = pd.read_excel(vaktliste_file_path, sheet_name=sheet_name)
#print(df.head(10))

#nye_kokker = 0 # 1: tar hensyn til nye kokker. 0: tar ikke hensyn til nye kokker


# Extract the dates into a list
date_format = '%Y-%m-%d %H:%M:%S' # Date format here must be equal to format in dataframe above
dates = df.iloc[0, 1:].dropna().tolist()
formatted_dates = [datetime.strptime(str(date), date_format).strftime('%d/%m ') for date in dates]
double_dates = [date for date in formatted_dates for _ in range(2)]
#print('Dates: ', double_dates)

#Extract the time slots
time_slots = df.iloc[1, 1::1].dropna().tolist()
compressed_time_slots = [time_slot[:2] + '-' + time_slot[8:10] for time_slot in time_slots]
#print('Time slots: ', compressed_time_slots)

#Create shift names
shift_names = []
for n in range(len(double_dates)):
    shift_names.append(double_dates[n] + compressed_time_slots[n])
#print(f'Shifts: {shift_names}')


# Extract relevant data from df and insert to another dataframe
start_of_chef_list = 3 
chefs_availability = df.iloc[start_of_chef_list:, :30].astype(object)
chefs_availability.columns = ['Kokk', 'Nyopptatt'] + shift_names
#print(chefs_availability.head(10))


hangarounds_row_index = chefs_availability[chefs_availability['Kokk'].str.contains("HANGAROUNDS", case=False, na=False)].index[0]

# Remove "AKTIVE" "HANGAROUNDS", "PANGER" AND NaN
chefs_availability = chefs_availability[
    ~chefs_availability['Kokk'].isin(["AKTIVE", "HANGAROUNDS", "PANGER"]) & 
    chefs_availability['Kokk'].notna()]

active_chefs = chefs_availability.iloc[:hangarounds_row_index-4, 0].dropna().tolist()
hangs_and_pangs = chefs_availability.iloc[hangarounds_row_index-4:, 0].dropna().tolist()
#print(f"Aktive kokker: {len(active_chefs)}{active_chefs}")
#print(f"Hangarounds og panger: {len(hangs_and_pangs)}{hangs_and_pangs}")



# List of chefs that has worked at least one semester
old_chefs = active_chefs.copy()

for chef in active_chefs:
    chef_index = chefs_availability[chefs_availability['Kokk'].str.contains(chef, case=False, na=False)].index[0]
    if chefs_availability.iloc[chef_index-3, 1] == 1.0:
        old_chefs.remove(chef)

chefs_availability = chefs_availability.drop(columns='Nyopptatt')

#If Chef have not doodled => available all shifts, except from hangarounds and pangs
for index, row in chefs_availability.iterrows():
    if index >= hangarounds_row_index:
        break
    if row[1:].isna().all():
        chefs_availability.loc[index, row.index[1:]] = "Kan jobbe!"

#print(chefs_availability[45:65])

doodled_chefs = active_chefs.copy()
for chef in hangs_and_pangs:
    chef_row = chefs_availability.loc[chefs_availability['Kokk'] == chef]
    if chef_row.iloc[:, 1:].eq("Kan jobbe!").any(axis=1).values[0]:
        doodled_chefs.append(chef)


week_1_df = chefs_availability.iloc[:, 0:number_of_shifts_week_1+1]
week_2_df = chefs_availability.iloc[:, [0] + list(range(number_of_shifts_week_1+1, number_of_shifts_week_1+number_of_shifts_week_2+1))]


# Returns a list for available chefs on shift 'dd/mm hh-hh'
def get_available_chefs(shift, temp_df):
    return temp_df[temp_df[shift] == "Kan jobbe!"]['Kokk'].tolist()

# Returns number of chefs needed for the shift
def get_num_of_chefs(shift_name):
    if '13-21' in shift_name:
        num_chefs = 3
    else:
        num_chefs = 4
    return num_chefs

# Check if there are at least one old chef at the shift
def have_enough_old_chefs(list_of_chefs):
    for chef in list_of_chefs:
        if chef in old_chefs:
            return True
        else:
            return False
        
# Creates and returns shift schedule for one week as a dataframe
def assign_chefs(availability_df):
    

    schedule = defaultdict(list) #Dictionary for the schedule

    #Sort the shifst by availability low to high
    df_without_chefs = availability_df.drop(columns=['Kokk'])
    kan_jobbe_count = df_without_chefs.apply(lambda col: col.value_counts().get("Kan jobbe!", 0))
    kan_jobbe_count_dict = kan_jobbe_count.to_dict()

    # Sort the dictionary by the number of "Kan jobbe!" in ascending order
    sorted_kan_jobbe_count = dict(sorted(kan_jobbe_count_dict.items(), key=lambda item: item[1], reverse=False))
    sorted_shifts = list(sorted_kan_jobbe_count.keys())

    # Print the sorted dictionary
    #print(sorted_shifts)


    temp_df = availability_df.copy()
    for shift in sorted_shifts:
        number_of_chefs = get_num_of_chefs(shift)
        available_chefs = get_available_chefs(shift, temp_df)
        enough_old_chefs = False
        cnt = 0
        while not enough_old_chefs: # Makes sure there is at least one old chef at the shift if possible
            random.shuffle(available_chefs)
            selected_chefs = available_chefs[:number_of_chefs]
            enough_old_chefs = have_enough_old_chefs(selected_chefs)
            cnt += 1
            if cnt > 10:
                break
        schedule[shift] = selected_chefs
        for chef in selected_chefs:
            for s in sorted_shifts:
                if s != shift:
                    temp_df.loc[temp_df['Kokk'] == chef, s] = "Opptatt"
                    # If hang/pang set as "Opptatt" in both weeks

    

    schedule_df = pd.DataFrame.from_dict(schedule, orient='index') # Load dictionary to dataframe
    schedule_df.columns = [f"Kokk {i+1}" for i in range(schedule_df.shape[1])] # Set column names

    

    # Reorganize dataframe chronologically
    schedule_df.index = pd.to_datetime(schedule_df.index, format='%d/%m %H-%M')
    schedule_sorted_df = schedule_df.sort_index()
    schedule_sorted_df.index = schedule_sorted_df.index.strftime('%d/%m %H-%M')#TING MÅ FIKSES HER9

    # Print the sorted DataFrame
    #print(schedule_sorted_df)

    return schedule_sorted_df

# Checks in schedule dataframe that all shifts are filled, no active chefs are excluded and that there are no duplicates
def check_schedule(availability_df):
    # Check for empty slots
    num_empty_slots = 10
    best_num_empty_slots = 20
    counter = 0
    max_iterations = 200
    while num_empty_slots >= 0 and counter < max_iterations:
        schedule_df = assign_chefs(availability_df)
        num_empty_slots = schedule_df.iloc[:,:3].isna().sum().sum()
        if num_empty_slots < best_num_empty_slots:
            final_schedule_df = schedule_df
            best_num_empty_slots = num_empty_slots
        counter += 1
    
    #print(f"Empty slots: {num_empty_slots}")

    
    assigned_chefs = final_schedule_df.values.ravel().tolist()
    assigned_chefs = [name for name in assigned_chefs if name != None]
    
    # Check for duplicates
    counts = Counter(assigned_chefs)
    duplicate_chefs = [item for item, count in counts.items() if count > 1]
    if len(duplicate_chefs) == 0:
        duplicate_chefs.append('None')

    # Check for excluded chefs
    excluded_chefs = []
    for chef in doodled_chefs:
        if chef not in assigned_chefs:
            excluded_chefs.append(chef)

    # Write this to excel file?
    #print(f'Excluded chefs: {len(excluded_chefs)}{excluded_chefs}')
    #print(f'Duplicated chefs: {duplicate_chefs}')

    return final_schedule_df, excluded_chefs


def add_excluded_chefs_to_schedule(schedule_df, excluded_chefs, availability_df):
    
    still_excluded_chefs = []

    chef_count_per_shift = schedule_df.notna().sum(axis=1) # Number of chefs per shift

    for chef in excluded_chefs:
        # Find the available shifts for the chef
        chef_row = availability_df.loc[availability_df['Kokk'] == chef]
        available_shifts = chef_row.iloc[:, 2:][chef_row.iloc[:, 2:] == "Kan jobbe!"].dropna(axis=1).columns.tolist()

        if available_shifts:  # If the chef has available shifts
            # Check the shifts with the least number of chefs and try to assign the chef
            for shift in available_shifts:
                min_chef_shift = chef_count_per_shift.idxmin()  # Find shift with the fewest chefs

                # If there's an empty spot in the shift, assign the chef
                if schedule_df.loc[min_chef_shift].isna().any():
                    first_empty_slot = schedule_df.loc[min_chef_shift].isna().idxmax()
                    schedule_df.loc[min_chef_shift, first_empty_slot] = chef

                    # Update the number of chefs in each shift after assignment
                    chef_count_per_shift = schedule_df.notna().sum(axis=1)
                    break  # Once chef is assigned, move to the next chef

        else:
            # If no available shifts, add chef to still_excluded_chefs list
            still_excluded_chefs.append(chef)

    # Return the updated schedule and the list of chefs that couldn't be assigned
    return schedule_df, still_excluded_chefs

    
# Save schedule to excel-file ab\nd make it pretty
def save_to_file(schedule_df,file_path, excluded_chefs):
    schedule_df.to_excel(file_path, index_label = file_path[27]+'. uke')
    workbook = load_workbook(file_path)
    worksheet = workbook.active

    column_widths = [11, 30, 30, 30, 30]  # Column widths
    for i, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[worksheet.cell(row=1, column=i).column_letter].width = width

    # Write in excluded chefs
    start_column = 'B'
    start_row = 18
    bold_font = Font(bold=True)
    worksheet[f"{start_column}{start_row}"].font = bold_font
    excluded_chefs.insert(0, "Ekskluderte kokker:")
    for i, chef in enumerate(excluded_chefs):
        cell_ref = f"{start_column}{start_row+i}"
        worksheet[cell_ref] = chef

    workbook.save(file_path)


def remove_scheduled_hangs_and_pangs(schedule_df):
    for chef in hangs_and_pangs:
        if chef in schedule_df.values:
            week_2_df.loc[week_2_df['Kokk'] == chef, week_2_df.columns[1:]] = "Opptatt"


# Write schedules to excel files
file_path_week_1 = 'test_data/Chef_Shifts_Week_1.xlsx'
file_path_week_2 = 'test_data/Chef_Shifts_Week_2.xlsx'

shift_schedule_week_1, excluded_chefs_1 = check_schedule(week_1_df)
final_schedule_week_1, final_excluded_chefs_1 = add_excluded_chefs_to_schedule(shift_schedule_week_1, excluded_chefs_1, week_1_df)
save_to_file(final_schedule_week_1, file_path_week_1, final_excluded_chefs_1)

remove_scheduled_hangs_and_pangs(shift_schedule_week_1)
#print(shift_schedule_week_1)
#print(week_2_df[52:])

shift_schedule_week_2, excluded_chefs_2 = check_schedule(week_2_df)
save_to_file(shift_schedule_week_2, file_path_week_2, excluded_chefs_2)
#print(shift_schedule_week_2)



print(f"Schedule saved to {file_path_week_1} and {file_path_week_2}")


#TODO
# Optimering: 1. Sørge for å ha minst to kokker på skift
# Håndtere ekskluderte kokker, som ikke er HELT opptatt

# Legge schedules inn i en fil, to sheets, samme format som sendt på mail
# Legge inn håndtering av ulike dato-formater i input fil


# Håndtering av vanlige errors
# Fikse slik at det fungerer med brøkdeler av uker
# Legge inn håndtering av 'Kan om nødvendig'
# Gjøre koden raskere. Implementere AC3?

#Noen nye kokker som mangler?